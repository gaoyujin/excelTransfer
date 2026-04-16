import os
import xlrd2
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from xlrd2 import XL_CELL_BLANK, XL_CELL_DATE, XL_CELL_EMPTY, XL_CELL_ERROR
from xlrd2.xldate import xldate_as_datetime

import common
import reChinese
from colorama import Fore, Style, init
import re

init(autoreset=True)


class FileName:
    pass


def process_string(input_str):
    """
    处理字符串中特定模式的子串：
    1. 以26个大写字母开头
    2. 后跟'_S '、'_S_'或'_S'
    将这些匹配的子串替换为空字符串
    """
    pattern = r'[A-Z]_S(?:_| |$)'
    return re.sub(pattern, '', input_str)


def _rgb_tuple_to_argb(rgb):
    """将 xlrd colour_map 的 RGB 元组转为 openpyxl ARGB 十六进制字符串。"""
    if not rgb:
        return None
    r, g, b = rgb
    return f"FF{r:02X}{g:02X}{b:02X}"


def _xlrd_indexed_colour(rb, colour_index):
    """从调色板解析颜色；无法解析时返回 None。"""
    if colour_index is None:
        return None
    return _rgb_tuple_to_argb(rb.colour_map.get(colour_index))


def _xlrd_line_style_to_side(style_idx, colour_argb):
    """xlrd 边框线型 → openpyxl Side。"""
    if not style_idx:
        return Side()
    style_map = {
        1: "thin",
        2: "medium",
        3: "dashed",
        4: "dotted",
        5: "thick",
        6: "double",
        7: "hair",
        8: "mediumDashed",
        9: "dashDot",
        10: "mediumDashDot",
        11: "dashDotDot",
        12: "mediumDashDotDot",
        13: "slantDashDot",
    }
    st = style_map.get(style_idx, "thin")
    return Side(style=st, color=colour_argb or "FF000000")


def _xlrd_font_to_openpyxl(rb, xf):
    """xlrd XF → openpyxl Font。"""
    fnt = rb.font_list[xf.font_index]
    color_argb = _xlrd_indexed_colour(rb, fnt.colour_index)
    is_bold = bool(fnt.bold)
    w = getattr(fnt, "weight", None)
    if w is not None and not is_bold:
        is_bold = w >= 600
    kwargs = {
        "name": fnt.name or "Calibri",
        "bold": is_bold,
        "italic": bool(fnt.italic),
    }
    if fnt.height:
        kwargs["size"] = fnt.height / 20.0
    if color_argb:
        kwargs["color"] = color_argb
    if fnt.struck_out:
        kwargs["strikethrough"] = True
    if fnt.underlined:
        kwargs["underline"] = "single"
    return Font(**kwargs)


def _xlrd_fill_to_openpyxl(rb, bg):
    """xlrd XFBackground → openpyxl PatternFill（无填充时返回 None）。"""
    if not bg.fill_pattern:
        return None
    fg = _xlrd_indexed_colour(rb, bg.pattern_colour_index)
    if fg:
        return PatternFill(fill_type="solid", fgColor=fg)
    bgc = _xlrd_indexed_colour(rb, bg.background_colour_index)
    if bgc:
        return PatternFill(fill_type="solid", fgColor=bgc)
    return None


def _xlrd_border_to_openpyxl(rb, xb):
    """xlrd XFBorder → openpyxl Border。"""
    def mk(side_style, side_colour_idx):
        c = _xlrd_indexed_colour(rb, side_colour_idx)
        return _xlrd_line_style_to_side(side_style, c)

    return Border(
        left=mk(xb.left_line_style, xb.left_colour_index),
        right=mk(xb.right_line_style, xb.right_colour_index),
        top=mk(xb.top_line_style, xb.top_colour_index),
        bottom=mk(xb.bottom_line_style, xb.bottom_colour_index),
    )


def _xlrd_alignment_to_openpyxl(al):
    """xlrd XFAlignment → openpyxl Alignment。"""
    hor_map = {0: "general", 1: "left", 2: "center", 3: "right", 4: "fill", 5: "justify", 6: "centerContinuous"}
    ver_map = {0: "top", 1: "center", 2: "bottom", 3: "justify", 4: "distributed"}
    return Alignment(
        horizontal=hor_map.get(al.hor_align, "general"),
        vertical=ver_map.get(al.vert_align, "bottom"),
        wrap_text=bool(al.text_wrapped),
        shrink_to_fit=bool(al.shrink_to_fit),
        indent=al.indent_level or 0,
    )


def _xlrd_cell_value(rb, sh, row, col):
    """读取单元格值（含日期）。"""
    cell = sh.cell(row, col)
    if cell.ctype in (XL_CELL_EMPTY, XL_CELL_BLANK):
        return None
    val = cell.value
    if cell.ctype == XL_CELL_DATE:
        try:
            val = xldate_as_datetime(val, rb.datemode)
        except Exception:
            pass
    elif cell.ctype == XL_CELL_ERROR:
        val = None
    return val


def _xlrd_merge_range_a1(rlo, rhi, clo, chi):
    """
    xlrd merged_cells 为半开区间 [rlo:rhi)、[clo:chi)，转为 A1 范围字符串。

    Returns:
        str: 例如 'B7:B8'。
    """
    start_row = rlo + 1
    end_row = rhi
    c1 = get_column_letter(clo + 1)
    c2 = get_column_letter(chi)
    if start_row == end_row and c1 == c2:
        return f"{c1}{start_row}"
    return f"{c1}{start_row}:{c2}{end_row}"


def _xls_to_openpyxl_with_styles(path):
    """
    将 .xls（BIFF）按 xlrd2 的格式信息转为 openpyxl 工作簿，保留字体、填充、边框、对齐、合并及行列尺寸。

    Args:
        path: .xls 文件路径。

    Returns:
        openpyxl.workbook.workbook.Workbook
    """
    rb = xlrd2.open_workbook(path, formatting_info=True)
    sh = rb.sheet_by_index(0)
    wb = Workbook()
    ws = wb.active
    ws.title = (sh.name or "Sheet1")[:31]

    for row in range(sh.nrows):
        for col in range(sh.ncols):
            cell = sh.cell(row, col)
            oc = ws.cell(row=row + 1, column=col + 1)
            val = _xlrd_cell_value(rb, sh, row, col)
            if val is not None:
                oc.value = val
            xf = rb.xf_list[cell.xf_index]
            oc.font = _xlrd_font_to_openpyxl(rb, xf)
            fill = _xlrd_fill_to_openpyxl(rb, xf.background)
            if fill is not None:
                oc.fill = fill
            oc.border = _xlrd_border_to_openpyxl(rb, xf.border)
            oc.alignment = _xlrd_alignment_to_openpyxl(xf.alignment)

    for rlo, rhi, clo, chi in sh.merged_cells:
        ws.merge_cells(_xlrd_merge_range_a1(rlo, rhi, clo, chi))

    for colx, info in sh.colinfo_map.items():
        if info.width:
            letter = get_column_letter(colx + 1)
            ws.column_dimensions[letter].width = info.width / 256.0

    for rowx, info in sh.rowinfo_map.items():
        if info.height:
            ws.row_dimensions[rowx + 1].height = info.height / 20.0

    return wb


def load_workbook_compat(path):
    """
    加载 BV 模板：.xlsx / .xlsm 使用 openpyxl（完整保留样式）；
    .xls 使用 xlrd2 带格式转换（保留顶部版式、合并单元格与行列尺寸）。

    openpyxl 无法直接读取二进制 .xls。

    Args:
        path: 模板文件路径。

    Returns:
        openpyxl.workbook.workbook.Workbook: 可供写入的工作簿。
    """
    lower = path.lower()
    if lower.endswith((".xlsx", ".xlsm")):
        return load_workbook(path)
    if lower.endswith(".xls"):
        try:
            return _xls_to_openpyxl_with_styles(path)
        except Exception:
            rb = xlrd2.open_workbook(path, formatting_info=False)
            sh = rb.sheet_by_index(0)
            wb = Workbook()
            ws = wb.active
            ws.title = (sh.name or "Sheet1")[:31]
            for row in range(sh.nrows):
                for col in range(sh.ncols):
                    cell = sh.cell(row, col)
                    if cell.ctype in (XL_CELL_EMPTY, XL_CELL_BLANK):
                        continue
                    val = cell.value
                    if cell.ctype == XL_CELL_DATE:
                        try:
                            val = xldate_as_datetime(val, rb.datemode)
                        except Exception:
                            pass
                    elif cell.ctype == XL_CELL_ERROR:
                        val = None
                    ws.cell(row=row + 1, column=col + 1, value=val)
            return wb
    return load_workbook(path)


def resolve_bv_template_workbook():
    """
    解析 BV 模板路径：优先使用 temp/drawing.xlsx（无损），否则使用 temp/drawing.xls。

    Returns:
        openpyxl.workbook.workbook.Workbook
    """
    base = "./temp/drawing"
    if os.path.exists(base + ".xlsx"):
        return load_workbook(base + ".xlsx")
    if os.path.exists(base + ".xlsm"):
        return load_workbook(base + ".xlsm")
    if os.path.exists(base + ".xls"):
        return load_workbook_compat(base + ".xls")
    raise FileNotFoundError(f"未找到 BV 模板：{base}.xlsx / .xlsm / .xls")


def process_filenames(filename):
    if (filename.endswith(".dwg") or filename.endswith(".DWG") or
        filename.endswith(".pdf") or filename.endswith(".PDF") or
        filename.endswith(".xls") or filename.endswith(".XLS") or
            filename.endswith(".xlsx") or filename.endswith(".XLSX")):
        # 查找第一个"SC"出现的位置
        sc_index = filename.find('SC')
        if sc_index == -1:
            sc_index = filename.find('sc')

        if sc_index != -1:
            new_name = filename[sc_index:]
            return new_name
        else:
            return filename
    else:
        return filename


def init():
    op_flag = True
    first = True
    while op_flag:
        try:

            if first:
                print(
                    f"{Fore.LIGHTGREEN_EX}{Style.BRIGHT}主人，把BV文件集合目录给我，我来收集图名、图号：{Style.RESET_ALL}")  # 打印蓝色文字
            else:
                print()
                print(f"{Fore.LIGHTGREEN_EX}{Style.BRIGHT}主人，奴仆我随时准备收集BV的图名、图号：{Style.RESET_ALL}")

            first = False
            file_path = input()

            if file_path == '切换':
                print()
                print(f"{Fore.CYAN}主人，处理程序开始切换，请根据提示操作...{Style.RESET_ALL}")
                op_flag = False
                raise KeyboardInterrupt

            # 路径不存在
            if not os.path.exists(file_path):
                print()
                print(
                    f"{Fore.RED}主人，我的能力不够，找不到这个路径，请您惩罚我吧！{Style.RESET_ALL}")
            else:
                print()
                print(f"{Fore.CYAN}奴仆正在努力劳作，主人您稍等......{Style.RESET_ALL}")
                print()
                # 读取文件夹中的所有文件
                file_list = os.listdir(file_path)

                if len(file_list) > 0:
                    last_data = []
                    set_data = []
                    count = 8

                    # 遍历文件
                    for xlsFile in file_list:
                        if (xlsFile.endswith(".dwg") or xlsFile.endswith(".DWG") or
                                xlsFile.endswith(".pdf") or xlsFile.endswith(".PDF") or
                                xlsFile.endswith(".xls") or xlsFile.endswith(".XLS") or
                                xlsFile.endswith(".xlsx") or xlsFile.endswith(".XLSX")):

                            new_xls_file = process_filenames(xlsFile)

                            file_name = common.get_file_name(new_xls_file)
                            names = file_name.split("_")

                            if len(names) < 1:
                                continue

                            if names[0] in set_data:
                                continue
                            else:
                                set_data.append(names[0])
                                my_data = FileName()
                                my_data.count = count
                                my_data.account = names[0]
                                my_data.type = names[1]
                                my_data.name = reChinese.remove_chinese_from_filename(process_string(file_name.replace(names[0] + '_', '')))
                                my_data.allName = new_xls_file

                                last_data.append(my_data)
                                count = count + 1

                    if len(last_data) > 0:
                        # 优先 drawing.xlsx；否则将 drawing.xls 按格式转为 openpyxl（保留顶部样式）
                        wb = resolve_bv_template_workbook()
                        # 根据索引获取指定sheet
                        sheet = wb.worksheets[0]

                        # 创建一个对象并设置属性
                        my_object = common.row_object(8, 8)

                        for item in last_data:
                            if not item or not item.account:
                                break

                            # 插入一行新的
                            common.copy_row_no_height(sheet, my_object, 1)

                            # 插入一行新的
                            row_index = item.count + 1
                            sheet['B' + str(row_index)] = item.account
                            sheet['D' + str(row_index)] = item.name
                            sheet['E' + str(row_index)] = item.type
                            sheet['G' + str(row_index)] = 'D'
                            sheet['H' + str(row_index)] = item.allName
                            my_object.rowIndex = my_object.rowIndex + 1

                        # 不存在 output 则创建
                        save_file_path = file_path + "\\bv_output"
                        common.create_directory_if_not_exists(save_file_path)

                        # openpyxl 输出为 OOXML（.xlsx），内容与 Excel 中另存为 xlsx 一致
                        save_path = file_path + "\\bv_output\\" + "BV_PostDrawings.xls"
                        wb.save(save_path)

                    print(f"{Fore.LIGHTBLACK_EX}{Style.BRIGHT}主人，收集BV图名、图号完成了：{file_path}{Style.RESET_ALL}")
                else:
                    print(
                        f"{Fore.RED}主人，您真好，路径下面是空的，是不是担心奴仆太累了？{Style.RESET_ALL}")
        except KeyboardInterrupt:
            raise KeyboardInterrupt

        except Exception as err:
            print(
                f"{Fore.RED}主人，不好了，奴仆的程序异常了，错误信息: ${err}{Style.RESET_ALL}")

